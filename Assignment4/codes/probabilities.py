#code to find theoretical values of probabilities, and write data to an excel file.

from openpyxl import load_workbook
import pandas as pd
import numpy as np

#function to find probabilities
def prob(x,list1):
    return list1[x]/sum(list1.values())

if __name__ == "__main__":

    wb = load_workbook("tables/raw_data.xlsx").active

#reading data

    items = [i for i in wb.iter_rows(min_row=1, max_row=1, values_only=True)]

    print(items)
    
    freq = [j for j in wb.iter_rows(min_row=2, max_row=2, values_only=True)]

    list1 = dict((x,y) for x,y in zip(items[0],freq[0]))
        
#creating list of probabilities
    problist = np.array([prob(i,list1) for i in range(10)])

#printing  theoretical probability value
    print("probability of occurrence of 6 is ",prob(6,list1))

#Sample size
    N = 2500

    counter = 0


#finding empirical value of probability, by counting frequency
    for i in range(N):
        x = np.random.choice([i for i in range(10)], p = problist)
        if(x == 6):
            counter = counter+1

#printing empirical value
    print("Experimental value of probability of occurrence of 6 is",counter/N)
    
